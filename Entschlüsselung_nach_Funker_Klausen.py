import openpyxl
import requests
from io import BytesIO

print('Nachricht verschlüsseln nach Funker Klausen\n')

abc = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', '.', '/']


while True:
    tabellen_bildung_eingabe = input('\nWort zur Bildung der Tabelle: ')
    
    if not tabellen_bildung_eingabe:
        print("\nKeine Eingabe erfolgt. Erneut versuchen!")
        continue

    tabellen_bildung = tabellen_bildung_eingabe.lower()

    if not all(char in abc for char in tabellen_bildung):
        print("\nDie Eingabe enthält ungültige Zeichen. Erneut versuchen!")
        continue
    
    if len(set(tabellen_bildung)) != len(tabellen_bildung):
        print("\nDie Eingabe enthält doppelte Zeichen. Erneut versuchen!")
        continue

    break



tabelle = [[0 for x in range(0)]for y in range(0)]

tabellen_bildung_liste = list(tabellen_bildung)

tabelle.append(tabellen_bildung_liste)




indexy = -1

unter_liste = []

for zeichen in abc:

    if zeichen not in tabellen_bildung_liste:

        unter_liste.append('')
        indexy += 1
        unter_liste [indexy] = zeichen

    if len(unter_liste) == len(tabellen_bildung_liste) or zeichen == '/':
        while len(unter_liste) < len(tabellen_bildung_liste):
            unter_liste.append('')


        tabelle.append(unter_liste)
        unter_liste = []
        indexy = -1


print ('\n\nTabelle:\n')

for y in range(len(tabelle)): 
    for x in range(len(tabellen_bildung_liste)):
        print (tabelle [y][x], end= ' ')
    print()

print()



asintoer= ['a','s','i','n','t','o','e','r']
zahlentabelle = [row.copy() for row in tabelle]



nummer = 0

for x in range(len(tabellen_bildung_liste)):
    for y in range(len(tabelle)):
    

    
        for zeichen in asintoer:
            if zahlentabelle [y][x] == zeichen:
                zahlentabelle [y][x] = nummer
                nummer += 1  
          
    





nummer = 80

for x in range(len(tabellen_bildung_liste)):
    for y in range(len(zahlentabelle)):
        if type(zahlentabelle [y][x]) == str:
            if zahlentabelle [y][x] != '':
                zahlentabelle [y][x] = nummer
                nummer += 1  





super=False

while True:

    if super:
        break

    while True:
        zu_enstschluesseln = input('\nZu entschlüsselnde Zahl: ')

        if not zu_enstschluesseln:
            print("\nKeine Eingabe erfolgt. Erneut versuchen!")
            continue

        if not zu_enstschluesseln.isdigit():
            print("\nDie Eingabe enthält ungültige Zeichen. Erneut versuchen!")
            continue
          
        if len(zu_enstschluesseln) < 10:
            print ("\nEingabe zu kurz. Mindestens zehn Zeichen eingeben.")
            continue
        else:
            break


    zu_entschluesseln_liste = list(zu_enstschluesseln)



    schluesselherkunft = zu_entschluesseln_liste[:5]  
    zu_entschluesseln_liste = zu_entschluesseln_liste[5:] 


    schluesselherkunft_entschluesselt = []

    for index in range(len(schluesselherkunft)):
        einerstelle = (int(schluesselherkunft [index]) - int(zu_entschluesseln_liste [index])) % 10
        schluesselherkunft_entschluesselt.append(einerstelle)

   




    seite = int(str(schluesselherkunft_entschluesselt[0])+str(schluesselherkunft_entschluesselt[1]))
    seite -= 1
    zeile = int(str(schluesselherkunft_entschluesselt[2])+str(schluesselherkunft_entschluesselt[3]))
    zeile -= 1
    spalte = int(schluesselherkunft_entschluesselt[4])
    spalte -= 1
    





    excel_url = 'https://github.com/fsgweimar/Verschluesselung_nach_Funker_Klausen/raw/main/chiffriertabelle.xlsx?raw=true'

    response = requests.get(excel_url)
    content = response.content

    excel_stream = BytesIO(content)


    chiffriertabelle = openpyxl.load_workbook(excel_stream, data_only=True)  # 'data_only=True' liest Formelergebnisse statt Formeln

    chiffriertabelle_aktiv = chiffriertabelle.active
    chiffriertabelle_liste = []



    try:

        sheet_name = chiffriertabelle.sheetnames[seite]

    except IndexError:
        print('\nEingabe ungültig. Erneut versuchen!')
        continue

    sheet = chiffriertabelle[sheet_name]

    for row in sheet.iter_rows(values_only=True):
        chiffriertabelle_liste.append(row)  




    chiffriertabelle.close()


    try:
        beginn_schluessel = list(str(chiffriertabelle_liste [zeile] [spalte]))
        
    except IndexError:
        print('Eingabe ungültig. Erneut versuchen!')
        continue




    letzte_zwei = beginn_schluessel  [-2:]
    beginn_schluessel = [letzte_zwei[0], letzte_zwei[1]]



    while len(beginn_schluessel) < len(zu_entschluesseln_liste):

        zeile+=1

        if zeile >= len(chiffriertabelle_liste):
            zeile = 0

        beginn_schluessel.extend(list(str(chiffriertabelle_liste [zeile] [spalte])))



    while len(beginn_schluessel) > len(zu_entschluesseln_liste):
        beginn_schluessel.pop()



    entschluesselt_1 = []

    for index in range(len(zu_entschluesseln_liste)):
        einerstelle = (int(zu_entschluesseln_liste [index]) - int(beginn_schluessel [index])) % 10
        entschluesselt_1.append(einerstelle)




    zahlen_liste=[]
    i = 0
    while i < len(entschluesselt_1):
        if entschluesselt_1[i] <= 7:
            zahlen_liste.append(entschluesselt_1[i])
            if i==len(entschluesselt_1)-1:
                super=True
            i += 1
        elif i+1<len(entschluesselt_1):
                anhaengen = int(str(entschluesselt_1[i])+str(entschluesselt_1[i + 1]))
                zahlen_liste.append(anhaengen)
                if i==len(entschluesselt_1)-2:
                    super=True
                i += 2
        else:
            print('\nEingabe ungültig. Erneut versuchen!')
            break
                

print ('\n\nNumerischer Klartext:\n')

for element in entschluesselt_1:
    print(element, end="")





entschluesselt_2=[]
index = 0

for zeichen in zahlen_liste:
    for y in range(len(zahlentabelle)):     
        for x in range(len(tabellen_bildung_liste)):
            if zahlentabelle [y][x] == zeichen:
                entschluesselt_2.append('')
                entschluesselt_2[index] = tabelle [y][x]
                index += 1




entschluesselt_2 = [str(element) for element in entschluesselt_2]
trennzeichen = ""
entschluesselt = trennzeichen.join(entschluesselt_2)

print('\n\n\nEntschlüsselt: \n')
print(entschluesselt)
