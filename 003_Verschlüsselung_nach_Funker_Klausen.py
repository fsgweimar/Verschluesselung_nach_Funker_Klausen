



import openpyxl

import requests
from io import BytesIO

print('\n\nNachricht verschlüsseln nach Funker Klausen\n')

abc = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', '.', '/']


while True:
    tabellen_bildung_eingabe = input('\nWort zur Bildung der Tabelle: ')
    
    if not tabellen_bildung_eingabe:
        print("\nKeine Eingabe erfolgt. Erneut versuchen!")
        continue

    tabellen_bildung = tabellen_bildung_eingabe.lower()

    if not all(char in abc for char in tabellen_bildung):
        print("\nDie Eingabe enthält ungültige Zeichen. Erneut versuchen!")
        continue
    
    if len(set(tabellen_bildung)) != len(tabellen_bildung):
        print("\nDie Eingabe enthält doppelte Zeichen. Erneut versuchen!")
        continue

    break




tabelle = [[0 for x in range(0)]for y in range(0)]

tabellen_bildung_liste = list(tabellen_bildung)

tabelle.append(tabellen_bildung_liste)




indexy = -1

unter_liste = []

for zeichen in abc:

    if zeichen not in tabellen_bildung_liste:

        unter_liste.append('')
        indexy += 1
        unter_liste [indexy] = zeichen

    if len(unter_liste) == len(tabellen_bildung_liste) or zeichen == '/':
        while len(unter_liste) < len(tabellen_bildung_liste):
            unter_liste.append('')


        tabelle.append(unter_liste)
        unter_liste = []
        indexy = -1


print ('\n\nTabelle:\n')

for y in range(len(tabelle)): 
    for x in range(len(tabellen_bildung_liste)):
        print (tabelle [y][x], end= ' ')
    print()

print()



asintoer= ['a','s','i','n','t','o','e','r']
zahlentabelle = [row.copy() for row in tabelle]



nummer = 0

for x in range(len(tabellen_bildung_liste)):
    for y in range(len(tabelle)):
    

    
        for zeichen in asintoer:
            if zahlentabelle [y][x] == zeichen:
                zahlentabelle [y][x] = nummer
                nummer += 1  
          
    





nummer = 80

for x in range(len(tabellen_bildung_liste)):
    for y in range(len(zahlentabelle)):
        if type(zahlentabelle [y][x]) == str:
            if zahlentabelle [y][x] != '':
                zahlentabelle [y][x] = nummer
                nummer += 1  

    
    

stop = 0

while True:
    zu_verschluesseln_eingabe = input('\nZu verschlüsselnder Text: ')

    if not zu_verschluesseln_eingabe:
        print("\nKeine Eingabe erfolgt. Erneut versuchen!")
        continue


    zu_verschluesseln = zu_verschluesseln_eingabe.lower()
    if not all(char in abc for char in zu_verschluesseln):
        print("\nDie Eingabe enthält ungültige Zeichen. Erneut versuchen!")
        continue

    if len(zu_verschluesseln) < 5:
        print ("\nEingabe zu kurz. Mindestens fünf Zeichen eingeben")
        continue
    else:
        break




zu_verschluesseln_liste = list(zu_verschluesseln)

 

verschlüsselt_1 = []
index = 0

for zeichen in zu_verschluesseln_liste:
    for y in range(len(tabelle)):     
        for x in range(len(tabellen_bildung_liste)):
            if tabelle [y][x] == zeichen:
                verschlüsselt_1.append('')
                verschlüsselt_1[index] = zahlentabelle [y][x]
                index += 1

print ('\n\nNumerischer Klartext:\n')

for element in verschlüsselt_1:
    print(element, end="")

print('\n\n')



excel_url = 'https://github.com/fsgweimar/Verschluesselung_nach_Funker_Klausen/raw/main/chiffriertabelle.xlsx?raw=true'

response = requests.get(excel_url)
content = response.content

excel_stream = BytesIO(content)


chiffriertabelle = openpyxl.load_workbook(excel_stream, data_only=True)  # 'data_only=True' liest Formelergebnisse statt Formeln

chiffriertabelle_aktiv = chiffriertabelle.active




chiffriertabelle_liste = []



while True:

    aufforderung= '\nNummer der Seite ('+str(len(chiffriertabelle.sheetnames))+' Seite(n) verfügbar): '
    seite_eingabe = input(aufforderung)
    

    if not seite_eingabe:
        print("\nKeine Eingabe erfolgt. Erneut versuchen!")
        continue

    if not seite_eingabe.isdigit():
        print("\nUngültige Eingabe. Zahl eingeben!")
        continue

    seite_eingabe = int (seite_eingabe)

    if seite_eingabe < 1:
        print("\nEingabe zu niedrig. Erneut Versuchen!")
        continue
    
    seite = seite_eingabe


    if seite <= len(chiffriertabelle.sheetnames):

        seite -= 1

        sheet_name = chiffriertabelle.sheetnames[seite]
        sheet = chiffriertabelle[sheet_name]

        chiffriertabelle_liste = []
        for row in sheet.iter_rows(values_only=True):
            chiffriertabelle_liste.append(row)        
        break
    else: 
        print("\nEingabe zu hoch. Erneut Versuchen!")



chiffriertabelle.close()


while True:

    aufforderung= '\nNummer der Zeile ('+str(len(chiffriertabelle_liste))+' Zeile(n) verfügbar): '
    zeile_eingabe = input(aufforderung)


    if not zeile_eingabe:
        print("\nKeine Eingabe erfolgt. Erneut versuchen!")
        continue

    if not zeile_eingabe.isdigit():
        print("\nUngültige Eingabe. Zahl eingeben!")
        continue

    zeile_eingabe = int (zeile_eingabe)

    if zeile_eingabe < 1:
        print("\nEingabe zu niedrig. Erneut Versuchen!")
        continue

    zeile = zeile_eingabe

    if zeile <= len(chiffriertabelle_liste):
        zeile -= 1
        break
    else:
        print("\nEingabe zu hoch. Erneut Versuchen!")




while True:

    aufforderung= '\nNummer der Spalte ('+str(len(chiffriertabelle_liste[zeile]))+' Spalte(n) verfügbar): '
    spalte_eingabe = input(aufforderung)


    if not spalte_eingabe:
        print("\nKeine Eingabe erfolgt. Erneut versuchen!")
        continue

    if not spalte_eingabe.isdigit():
        print("\nUngültige Eingabe. Zahl eingeben!")
        continue

    spalte_eingabe = int (spalte_eingabe)

    if spalte_eingabe < 1:
        print("\nEingabe zu niedrig. Erneut Versuchen!")
        continue

    spalte = spalte_eingabe

    if spalte <= len(chiffriertabelle_liste[zeile]):
        spalte -= 1
        break
    else:
        print("\nEingabe zu hoch. Erneut Versuchen!")




schluessel = [] 

beginn_schluessel = list(str(chiffriertabelle_liste [zeile] [spalte]))

letzte_zwei = beginn_schluessel  [-2:]
beginn_schluessel = [letzte_zwei[0], letzte_zwei[1]]


verschlüsselt_1 = [str(element) for element in verschlüsselt_1]
trennzeichen = ""
verschlüsselt_1 = trennzeichen.join(verschlüsselt_1)


verschlüsselt_1 = list(verschlüsselt_1)



while len(beginn_schluessel) < len(verschlüsselt_1):

    zeile+=1

    if zeile >= len(chiffriertabelle_liste):
        zeile = 0

    beginn_schluessel.extend(list(str(chiffriertabelle_liste [zeile] [spalte])))







while len(beginn_schluessel) > len(verschlüsselt_1):
    beginn_schluessel.pop()



verschlüsselt_2 = []

for index in range(len(verschlüsselt_1)):
    einerstelle = (int(verschlüsselt_1 [index]) + int(beginn_schluessel [index])) % 10
    verschlüsselt_2.append(einerstelle)





seite_zehner = (seite_eingabe // 10) % 10
seite_einer = seite_eingabe % 10 

zeile_zehner = (zeile_eingabe // 10) % 10
zeile_einer = zeile_eingabe % 10

spalte_einer = spalte_eingabe % 10


schlüsselherkunft = str(seite_zehner) + str(seite_einer) + str(zeile_zehner) + str(zeile_einer) + str(spalte_einer)



schlüsselherkunft = list(schlüsselherkunft)



schlüsselherkunft_verschlüsselt = []

for index in range(len(schlüsselherkunft)):
    einerstelle = (int(schlüsselherkunft [index]) + int(verschlüsselt_2 [index])) % 10
    schlüsselherkunft_verschlüsselt.append(einerstelle)



verschlüsselt_liste = schlüsselherkunft_verschlüsselt

verschlüsselt_liste.extend(verschlüsselt_2)




verschlüsselt_liste = [str(element) for element in verschlüsselt_liste]
trennzeichen = ""
verschlüsselt = trennzeichen.join(verschlüsselt_liste)

print ('\n\nVerschlüsselt: \n')
print (verschlüsselt)



